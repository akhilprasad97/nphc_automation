import os
from openpyxl import load_workbook
from datetime import datetime
 

def appendData(scriptPath,empID,customerName,platformName,fileName):
     os.chdir(scriptPath)
     timeStamp = datetime.now()
     wb = load_workbook("usageDetails.xlsx")
     sheet = wb["usageDetails"]
     maxrow = sheet.max_row
     sheet.cell(row = maxrow + 1,column = 1).value = empID
     sheet.cell(row = maxrow + 1,column = 2).value = customerName
     sheet.cell(row = maxrow + 1,column = 3).value = platformName
     sheet.cell(row = maxrow + 1,column = 4).value = fileName
     sheet.cell(row = maxrow + 1,column = 5).value = timeStamp
     sheet.cell(row = maxrow + 1,column = 6).value = maxrow + 1 
     sheet.cell(row = maxrow + 1,column = 7).value = "Not started"
     wb.save("usageDetails.xlsx")
